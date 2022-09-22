# -*- encoding=utf8 -*-

from airtest.core.api import *
from airtest.core.android import Android

import openpyxl
import random
import Basic
import Case
import Const


class AirTest(Basic.Steps, Case.Flow):

    def __init__(self, device_=None, speed=1):
        super().__init__(speed)
        android = Android()
        # 截图像素精度75
        ST.SNAPSHOT_QUALITY = 75
        self.width = android.display_info["width"]
        self.height = android.display_info["height"]
        self.resolution = (self.width, self.height)
        # 这里把一个对照的设备4个装备绝对坐标，再计算出根据屏幕分辨率计算出来的相对路径，会有10以内的像素偏差
        self.EQUIP1_A_POS = (round(150 / 1224, 2) * self.width, round(930 / 2700, 2) * self.height)
        self.EQUIP2_A_POS = (round(1100 / 1224, 2) * self.width, round(930 / 2700, 2) * self.height)
        self.EQUIP3_A_POS = (round(150 / 1224, 2) * self.width, round(1130 / 2700, 2) * self.height)
        self.EQUIP4_A_POS = (round(1100 / 1224, 2) * self.width, round(1130 / 2700, 2) * self.height)
        # 获得配置
        filepath = "./ResourceConfig.xlsx"
        workbook = openpyxl.load_workbook(filepath)
        # steps资源
        sheet = workbook.get_sheet_by_name("steps")
        max_row = sheet.max_row + 1
        self.steps = list()
        for i in range(2, max_row):
            step = dict()
            step["template_name"] = sheet[f"B{i}"].value
            step["img_name"] = f"img/{sheet[f'C{i}'].value}.png"
            step["record_pos_x"] = sheet[f"D{i}"].value
            step["record_pos_y"] = sheet[f"E{i}"].value
            self.steps.append(step)
        # hero资源
        sheet = workbook.get_sheet_by_name("hero")
        max_row = sheet.max_row + 1
        self.hero = list()
        for i in range(2, max_row):
            h = dict()
            h["template_name"] = sheet[f"B{i}"].value
            h["img_name"] = f"img/hero/{sheet[f'C{i}'].value}.png"
            self.hero.append(h)
        # fetter资源
        sheet = workbook.get_sheet_by_name("fetter")
        max_row = sheet.max_row + 1
        self.fetters = list()
        for i in range(2, max_row):
            fetter = dict()
            fetter["template_name"] = sheet[f"B{i}"].value
            fetter["img_name"] = f"img/fetter/{sheet[f'C{i}'].value}.png"
            fetter["hero"] = sheet[f"D{i}"].value
            self.fetters.append(fetter)
        # 连接设备
        auto_setup(__file__, devices=[f"Android://127.0.0.1:5037/{device_}?cap_method=JAVACAP&&ori_method=MINICAPORI&&touch_method=MAXTOUCH"], logdir=Const.LOG_DIR)

    def easy_touch(self, template):
        touch_obj = self.get_template(template)
        if exists(touch_obj):
            touch(touch_obj)
        sleep(self.speed)

    def wait_to_touch(self, template, timeout=120):
        """
        在超时时间内等待对象出现并进行点击，超过时间未出现则抛出异常

        :param template: 图片资源
        :param timeout: 超时时间
        """
        obj = self.get_template(template)
        try:
            wait(obj, timeout=timeout)
            touch(obj)
        except TargetNotFoundError:
            raise AssertionError("%s doesn't exist in screen" % template)

    def wait_for_any_to_touch(self, templates, timeout=120):
        """
        在超时时间内等待对象出现并进行点击，超过时间未出现则抛出异常

        :param templates: 图片资源，可以传递一个对象数组
        :param timeout: 超时时间
        """
        start = time.time()
        while 1:
            for template in templates:
                obj = self.get_template(template)
                if exists(obj):
                    touch(obj)
                    sleep(self.speed)
                    return
            if time.time() - start > timeout:
                raise AssertionError("no template appear")
            sleep(1.44)

    def wait_show_up(self, template, timeout=30):
        """
        在超时时间内等待对象出现，超过时间未出现则抛出异常

        :param template:
        :param timeout:
        """
        obj = self.get_template(template)
        try:
            wait(obj, timeout=timeout)
        except TargetNotFoundError:
            raise AssertionError("%s doesn't exist in screen" % template)
        finally:
            sleep(self.speed)

    def loop_wait(self, template):
        obj = self.get_template(template)
        if exists(obj):
            touch(obj)
        else:
            self.loop_wait(template)
        sleep(self.speed)

    def swipe_loop_find(self, template, length, direction=True) -> bool:
        """
        持续滑动查找指定对象，可指定滑动的方向、幅度，并且在超过了一定的滑动次数后没找到当前对象退出

        :param template: 资源对象
        :param length: 滑动幅度 （选择横向滑动时幅度应该小一点）
        :param direction: 滑动方向 True：竖向 False：横向
        :return: 返回对象是否存在 True：存在 False：不存在
        """
        swipe_count = 0
        break_count = 0
        center_x = self.width / 2
        center_y = self.height / 2
        while 1:
            if break_count == 50:
                # 这里是为了防止在不可以滑动的页面，到达一定次数会退出这个死循环
                return False
            if exists(template):
                return True
            else:
                break_count += 1
                if direction:
                    # 竖向滑动
                    if swipe_count < 5:
                        start_pos = (center_x, center_y + 200)
                        end_pos = (center_x, center_y - length)
                        swipe(start_pos, end_pos)
                        swipe_count += 1
                    if swipe_count >= 5:
                        start_pos = (center_x, center_y - 200)
                        end_pos = (center_x, center_y + length)
                        swipe(start_pos, end_pos)
                        swipe_count += 1
                    if swipe_count >= 10:
                        swipe_count = 0
                else:
                    # 横向滑动
                    if swipe_count < 5:
                        start_pos = (center_x + 100, center_y)
                        end_pos = (center_x - length, center_y)
                        swipe(start_pos, end_pos)
                        swipe_count += 1
                    if swipe_count > 5:
                        start_pos = (center_x - 100, center_y)
                        end_pos = (center_x + length, center_y)
                        swipe(start_pos, end_pos)
                        swipe_count += 1
                    if swipe_count >= 10:
                        swipe_count = 0

    def get_template(self, name, rgb=False):
        for step in self.steps:
            if name == step.get("template_name"):
                img_name = step.get("img_name")
                record_pos_x = step.get("record_pos_x")
                record_pos_y = step.get("record_pos_y")
                if rgb:
                    return Template(rf"{img_name}", record_pos=(record_pos_x, record_pos_y), resolution=self.resolution, rgb=True)
                else:
                    return Template(rf"{img_name}", record_pos=(record_pos_x, record_pos_y), resolution=self.resolution)
        raise Exception("资源不存在", name)

    def get_hero(self, name):
        for h in self.hero:
            if name == h.get("template_name"):
                img_name = h.get("img_name")
                return Template(rf"{img_name}", resolution=self.resolution)
        raise Exception("英雄不存在", name)

    def get_fetter(self, name):
        for fetter in self.fetters:
            if name == fetter.get("template_name"):
                img_name = fetter.get("img_name")
                return Template(rf"{img_name}", resolution=self.resolution)
        raise Exception("羁绊不存在", name)

    def solider(self, name):
        if name in ["步兵", "盾甲兵", "重盾兵"]:
            return Template(r"./img/solider/tpl1652432199077.png", record_pos=(-0.45, -0.878), resolution=self.resolution)
        if name in ["弓兵", "长弓兵", "神射手"]:
            return Template(r"./img/solider/tpl1652432589636.png", record_pos=(-0.267, -0.878), resolution=self.resolution)
        if name in ["枪兵", "长枪兵", "长戟兵"]:
            return Template(r"./img/solider/tpl1652432614354.png", record_pos=(-0.082, -0.876), resolution=self.resolution)
        if name in ["骑兵", "铁骑兵", "重骑兵"]:
            return Template(r"./img/solider/tpl1652432625426.png", record_pos=(0.104, -0.875), resolution=self.resolution)
        if name in ["木质龙雀", "银甲龙雀", "金灵龙雀"]:
            return Template(r"./img/solider/tpl1652432634803.png", record_pos=(0.287, -0.877), resolution=self.resolution)

    def begin_fight(self):
        self.easy_touch("FightBtn")

    def start_fight(self):
        self.loop_wait("FightBtn")

    def next_fight(self):
        self.loop_wait("NextFightBtn")

    def choose_buff(self):
        if exists(self.get_template("ChooseBuffTitle")):
            index = random.randint(1, 4)
            if index == 1:
                x = int(0.2 * self.width)
                y = int(0.5 * self.height)
                touch((x, y))
            if index == 2:
                x = int(0.5 * self.width)
                y = int(0.5 * self.height)
                touch((x, y))
            if index == 3:
                x = int(0.8 * self.width)
                y = int(0.5 * self.height)
                touch((x, y))
            if index == 4:
                # 全选
                self.easy_touch("ChooseAllBuffBtn")
        sleep(self.speed)

    def resurgence(self):
        self.easy_touch("ResurgenceBtn")

    def try_fight_again(self):
        self.easy_touch("FightAgainBtn")

    def hero_tab(self):
        self.easy_touch("HeroTabBtn")

    def choose_hero(self, name):
        hero = self.get_hero(name)
        if self.swipe_loop_find(hero, 500):
            touch(hero)
        sleep(self.speed)

    def choose_solider(self, name):
        solider = self.solider(name)
        if exists(solider):
            touch(solider)
        sleep(self.speed)

    def lv_upgrade(self, duration=0.01):
        upgrade_btn = self.get_template("UpgradeBtn")
        if exists(upgrade_btn):
            touch(upgrade_btn, duration=duration)
        sleep(self.speed)

    def star_up(self):
        self.easy_touch("StarUpBtn")

    def go_star_up_btn(self):
        self.easy_touch("GoStarUpBtn")

    def suit_up(self):
        self.easy_touch("SuitUpBtn")

    def take_off(self):
        self.easy_touch("TakeOffBtn")

    def gift_tab(self):
        self.easy_touch("GiftTabBtn")

    def train(self, duration=0.01):
        train_btn = self.get_template("TrainBtn")
        if exists(train_btn):
            touch(train_btn, duration=duration)
        sleep(self.speed)

    def fast_buy_toggle(self, fast_buy=False):
        switch_on = self.get_template("FastBuyOn")
        switch_off = self.get_template("FastBuyOn")
        if fast_buy:
            if exists(switch_off):
                touch(switch_off)
        else:
            if exists(switch_on):
                touch(switch_on)
        sleep(self.speed)

    def main_tab(self):
        self.easy_touch("MainTabBtn")

    def back_btn(self):
        self.easy_touch("BackBtn")

    def get_reward(self):
        self.loop_wait("GetRewardBtn")

    def manor_tab(self):
        self.easy_touch("ManorTabBtn")

    def recruit(self, is_manor_btn=True, recruit_times=1, is_skip_animate=True):
        if is_manor_btn:
            self.easy_touch("RecruitEntryBtn")
        if is_skip_animate:
            self.easy_touch("SkipAnimateBtn")
        if recruit_times != 1:
            self.easy_touch("TenTimesWheelBtn")
        else:
            once_btn = self.get_template("OnceWheelBtn")
            free_btn = self.get_template("FreeWheelBtn")
            if exists(once_btn):
                touch(once_btn)
            elif exists(free_btn):
                touch(free_btn)
        sleep(self.speed)

    def one_key_add_piece(self):
        self.easy_touch("OneKeyAddBtn")

    def confirm(self):
        self.easy_touch("ConfirmBtn")

    def click_close(self):
        self.easy_touch("ClickClose")

    def go_break(self):
        self.easy_touch("GoBreakBtn")

    def go_break2(self):
        self.easy_touch("GoBreakBtn2")

    def break_(self):
        self.easy_touch("BreakBtn")

    def close(self):
        self.easy_touch("CloseBtn")

    def hero_task(self):
        self.easy_touch("HeroTaskBtn")

    def get_task_reward(self):
        get_task_reward_btn = self.get_template("GetHeroTaskRewardBtn")
        if exists(get_task_reward_btn):
            touch(get_task_reward_btn)
            self.get_task_reward()
        sleep(self.speed)

    def choose_equip(self, index):
        if index == 1:
            touch(self.EQUIP1_A_POS)
        if index == 2:
            touch(self.EQUIP2_A_POS)
        if index == 3:
            touch(self.EQUIP3_A_POS)
        if index == 4:
            touch(self.EQUIP4_A_POS)
        sleep(self.speed)

    def one_key_add_equip_piece(self):
        self.easy_touch("OneKeyAddEquipPieceBtn")

    def equip_lv_up(self):
        self.easy_touch("EquipUpgradeBtn")

    def go_equip_upgrade(self):
        self.easy_touch("GoEquipUpgradeBtn")

    def go_equip_star_up(self):
        self.easy_touch("GoEquipStarUpBtn")

    def equip_star_up(self):
        self.easy_touch("EquipStarUpBtn")

    def fetter_btn(self):
        self.easy_touch("FetterBtn")

    def choose_fetter(self, name):
        fetter = self.get_fetter(name)
        if self.swipe_loop_find(fetter, 600):
            touch(fetter)
        sleep(self.speed)

    def army_btn(self):
        self.easy_touch("ArmyBtn")

    def hero_buff_btn(self):
        self.easy_touch("HeroBuffBtn")

    def fetter_upgrade_btn(self):
        self.easy_touch("FetterUpgradeBtn")

    def risk_btn(self):
        self.easy_touch("RiskBtn")

    def mhy_entry(self):
        self.easy_touch("MHYEntry")

    def mhy_tips(self):
        self.easy_touch("MHYTips")

    def mhy_challenge_btn(self):
        self.easy_touch("MHYChallengeBtn")

    def mhy_fight_result_get_btn(self):
        self.easy_touch("MHYFightResultGetBtn")

    def newbie_guide(self):
        """
        新手引导
        """
        def buy_solider_bow(times=1):
            btn = Template(r"./img/tpl1652784840660.png", record_pos=(0.261, 0.953), resolution=self.resolution)
            if exists(btn):
                touch(btn)
            sleep(self.speed)
            btn = Template(r"./img/tpl1652785066536.png", record_pos=(-0.196, 0.686), resolution=self.resolution)
            while times > 0:
                if exists(btn):
                    touch(btn)
                    sleep(2)
                    times -= 1
            sleep(self.speed)

        def release_skill():
            btn = Template(r"./img/tpl1652786160536.png", record_pos=(-0.002, 0.912), resolution=self.resolution)
            while 1:
                if exists(btn):
                    touch(btn)
                if exists(self.get_template("NextFightBtn")):
                    break
                sleep(1)

        self.wait_to_touch("GetSoldierBtn_NG")
        self.wait_to_touch("FightBtn")
        self.wait_to_touch("GetRewardBtn", timeout=60)
        sleep(2)
        swipe((400, 1300), (600, 1300), duration=1)
        sleep(2)
        self.wait_to_touch("GetSoldierBtn_NG")
        sleep(1)
        self.wait_to_touch("FightBtn")
        self.wait_to_touch("NextFightBtn", timeout=60)
        self.wait_to_touch("GetBtn")
        buy_solider_bow(times=2)
        count = 2
        while count > 0:
            self.wait_to_touch("FightBtn")
            self.wait_to_touch("NextFightBtn", timeout=60)
            count -= 1
        buy_solider_bow(times=5)
        self.wait_to_touch("FightBtn")
        self.wait_to_touch("GetRewardBtn", timeout=60)
        self.wait_to_touch("GetBtn")
        # 这里有个动画需要sleep
        sleep(5)
        self.hero_tab()
        self.wait_show_up("UpgradeBtn")
        self.lv_upgrade(duration=3)
        self.wait_to_touch("BackBtn")
        self.wait_to_touch("MainTabBtn")
        self.wait_to_touch("FightBtn")
        self.wait_to_touch("FightBtn")
        release_skill()
        count = 4
        while count > 0:
            self.wait_to_touch("NextFightBtn", timeout=60)
            self.wait_to_touch("FightBtn")
            count -= 1
        self.wait_to_touch("GetRewardBtn", timeout=60)
        sleep(3)
        self.manor_tab()
        self.recruit(is_skip_animate=False)
        pass

    def upgrade(self, names, get_task_reward=False):
        """
        英雄升级，可以领取升级任务奖励

        :param names: 英雄名称
        :param get_task_reward: 是否需要领取英雄升级任务奖励
        """
        self.hero_tab()
        for name in names:
            self.choose_hero(name)
            tiger_trigger = False
            # 试探性点击
            self.lv_upgrade()
            while 1:
                if tiger_trigger:
                    if exists(self.get_template("TigerTag")):
                        self.click_close()
                        self.click_close()
                        self.fast_buy_toggle(True)
                        tiger_trigger = False
                if exists(self.get_template("GoStarUpPanel")) or exists(self.get_template("StarUpChoosePiecePanel")):
                    self.star_up()
                    self.go_star_up_btn()
                    if exists(self.get_template("StarUpChoosePiecePanel")):
                        self.one_key_add_piece()
                        self.confirm()
                        if exists(self.get_template("WarmPrompt")):
                            btn1 = self.get_template("WarmPromptConfirm")
                            if exists(btn1):
                                touch(btn1)
                        sleep(self.speed)
                        self.click_close()
                elif exists(self.get_template("GoBreakBtn")):
                    self.go_break()
                    self.go_break2()
                    while 1:
                        self.break_()
                        wait(self.get_template("BreakInfoPanel"), timeout=10)
                        self.click_close()
                        if not exists(self.get_template("BreakBtn", rgb=True)) and exists(self.get_template("BreakLevelTopTag", rgb=True)):
                            self.back_btn()
                            break
                if exists(self.get_template("HeroMaxLevelTag", rgb=True)):
                    if get_task_reward:
                        self.hero_task()
                        self.get_task_reward()
                        self.click_close()
                    self.back_btn()
                    break
                self.lv_upgrade(duration=5)

    def equip_upgrade(self, name, equip_index, one_key_max=True, current_level=1, level=1, suit_up=True):
        """
        装备升级

        :param name: 英雄名称
        :param equip_index: 升级某个装备或者全部升级，单个装备升级传入1-4，全部升级请传入0
        :param one_key_max: 一键把装备升满级，这个参数为True的话，current_level & level这两个参数不生效，二选其一
        :param current_level: 装备当前等级，这个参数能极大提升运行效率
        :param level: 对指定装备提升多少等级
        :param suit_up: 是否需要一键穿戴装备，默认会一键点击穿戴装备
        """

        def fill_piece():
            pos = (205, 1080)
            touch(pos, duration=2)
            self.equip_lv_up()

        def normal_upgrade():
            nonlocal count
            while range(0, level):
                self.one_key_add_equip_piece()
                self.equip_lv_up()
                count += 1
                if level == count and (count + current_level) == 50:
                    break

        self.hero_tab()
        self.choose_hero(name)
        count = 1
        if suit_up:
            self.suit_up()
            # 防止一键上装备的弹窗
            self.click_close()
            self.click_close()
        if 1 <= equip_index <= 4:
            self.choose_equip(equip_index)
            self.go_equip_upgrade()
            if one_key_max:
                fill_piece()
            else:
                normal_upgrade()
            self.back_btn()
        if equip_index == 0:
            for index in range(1, 5):
                self.choose_equip(index)
                self.go_equip_upgrade()
                if one_key_max:
                    fill_piece()
                else:
                    normal_upgrade()
                self.back_btn()

    def fetter_upgrade(self, name):
        """
        羁绊激活升级，需要先激活当前羁绊，如条件不满足时会将所对应的羁绊的英雄升到顶级

        :param name: 羁绊名称
        """
        self.hero_tab()
        self.fetter_btn()
        self.choose_fetter(name)
        gray_fetter_upgrade_btn = self.get_template("GrayFetterUpgradeBtn", rgb=True)
        fetter_upgrade_btn = self.get_template("FetterUpgradeBtn", rgb=True)
        skill_active_btn = self.get_template("FetterSkillActiveBtn")
        if exists(skill_active_btn):
            # 有时候会进入到武神技能界面，需要跳转到羁绊升级页面
            self.hero_buff_btn()
        if exists(gray_fetter_upgrade_btn) and not exists(fetter_upgrade_btn):
            # 需要升级英雄才能升级羁绊，跳转到英雄升级
            print("升级英雄")
            self.back_btn()
            self.army_btn()
            for fetter in self.fetters:
                if name == fetter["template_name"]:
                    hero = fetter["hero"].split("|")
                    self.upgrade(hero)
            self.fetter_btn()
            self.choose_fetter(name)
        while 1:
            self.fetter_upgrade_btn()
            self.click_close()
            if not exists(gray_fetter_upgrade_btn) and not exists(fetter_upgrade_btn):
                self.back_btn()
                self.main_tab()
                break

    def mhy(self, times):
        self.risk_btn()
        self.mhy_entry()
        self.mhy_tips()
        while times:
            self.mhy_challenge_btn()
            wait(self.get_template("FightBtn"), timeout=10)
            self.begin_fight()
            self.choose_buff()
            wait(self.get_template("MHYFightResultGetBtn"), timeout=120)
            self.mhy_fight_result_get_btn()
            sleep(8)
            if exists(self.get_template("NewRecord")):
                touch((self.width / 2, self.height / 2))
                self.close()
            times -= 1
        pass

    def infinity_war(self, current_floor, fight_times):
        self.begin_fight()
        while fight_times:
            self.wait_to_touch("FightBtn", timeout=240)
            if current_floor % 5 == 0:
                self.choose_buff()
            self.wait_for_any_to_touch(["NextFightBtn", "GetRewardBtn"], timeout=240)
            sleep(5)
            current_floor += 1
            fight_times -= 1

    def random_play(self, index):
        while 1:
            if index == 1:
                self.infinity_war(1, 10)
                self.back_btn()
            # 回到首页
            self.upgrade([])
            # 回到首页
            self.mhy(10)
            # 回到首页
            self.recruit()
            # 回到首页
        pass
